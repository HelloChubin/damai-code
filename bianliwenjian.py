# -*-coding:utf-8-*- 
print("--------------------------------------------！")
import os
import openpyxl
from openpyxl.styles import PatternFill

# files = list();
def traverseFile(pathName): 
    if os.path.exists(pathName):
        fileList = os.listdir(pathName)
        print(type(fileList))
        for fileName in fileList:
            file_name, file_end = os.path.splitext(fileName)
            if os.path.splitext(fileName)[1] == ".xlsx":
                # print(os.path.join(pathName,fileName))
                finallyname = (os.path.join(pathName,fileName))
                try:
                    wb = openpyxl.load_workbook(finallyname)
                    # wbSheet = wb.worksheets[0]
                    ws = wb.active
                    a1 = ws['A1']
                    # 3-设置样式，并且加载到对应单元格
                    fill = PatternFill("solid", fgColor="1874CD")
                    a1.fill = fill
                    wb.save(finallyname)
                    wb.close()
                    # wb.save(finallyname)
                    print("已经打开过："+finallyname)
                except:
                    print("报错了，请检查{}文件是否打开着".format(finallyname))

traverseFile(r"C:\Users\Administrator\Documents\系统文档")

